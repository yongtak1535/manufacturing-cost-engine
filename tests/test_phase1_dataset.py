from collections import Counter
from decimal import Decimal
from pathlib import Path

from manufacturing_cost_engine.loader import (
    excel_files,
    load_dataset,
    duplicate_file_groups,
)
from manufacturing_cost_engine.validation import (
    validate_duplicate_files, validate_period_rows, validate_work_orders,
    validate_material_issues, validate_bom_issues, validate_bom_version,
    validate_tolerance_rules, validate_routing, validate_labor,
    validate_labor_hours, validate_gl_balance, validate_gl_period,
    validate_account_mapping, validate_standard_cost, validate_actual_cost,
    validate_gl_reconciliation, validate_contract, validate_direct_expense,
)
from manufacturing_cost_engine.cost_engine import (
    calculate_actual_total_cost_by_wo, calculate_total_variance_by_wo,
)


DATASET_DIR = Path("hanbit_mvp_dataset_phase1")


def test_phase1_dataset_directory_exists():
    assert DATASET_DIR.exists()
    assert DATASET_DIR.is_dir()


def test_phase1_core_files_exist():
    required_files = [
        "01_company_plant.xlsx",
        "02_period.xlsx",
        "03_cost_center.xlsx",
        "04_cost_element.xlsx",
        "20_work_order.xlsx",
        "21_production_output.xlsx",
        "22_material_issue.xlsx",
    ]

    for filename in required_files:
        assert (DATASET_DIR / filename).exists(), filename


def test_excel_files_exclude_expected_result_files():
    files = excel_files(DATASET_DIR)

    names = [path.name for path in files]

    assert "90_expected_results.xlsx" not in names
    assert "91_error_catalog.xlsx" not in names


def test_phase1_dataset_can_be_loaded():
    data = load_dataset(DATASET_DIR)

    assert data

    assert any(
        key.startswith("01_company_plant.xlsx::")
        for key in data
    )

    assert any(
        key.startswith("02_period.xlsx::")
        for key in data
    )

    assert any(
        key.startswith("20_work_order.xlsx::")
        for key in data
    )

    assert any(
        key.startswith("22_material_issue.xlsx::")
        for key in data
    )


def test_phase1_loaded_rows_have_source_row():
    data = load_dataset(DATASET_DIR)

    loaded_rows = [
        row
        for rows in data.values()
        for row in rows
    ]

    assert loaded_rows

    assert all(
        "_source_row" in row
        for row in loaded_rows
    )
def test_phase1_dataset_has_expected_scale():
    data = load_dataset(DATASET_DIR)

    # 25개 로딩 대상 Excel (Phase 2: 30_contract.xlsx, 31_direct_expense.xlsx,
    # 34_direct_expense_budget.xlsx, 32_cost_rate_rule.xlsx,
    # contract_material_supply_type.xlsx 추가)
    files = excel_files(DATASET_DIR)

    assert len(files) == 25

    # 주요 거래 데이터가 실제로 로드되었는지 확인
    assert len(data["20_work_order.xlsx::work_order"]) == 20
    assert len(data["21_production_output.xlsx::production_output"]) == 19
    assert len(data["22_material_issue.xlsx::material_issue"]) == 62
    assert len(data["23_labor_transaction.xlsx::labor_transaction"]) == 54
    assert len(data["24_gl_transaction.xlsx::gl_transaction"]) == 44


def test_phase1_dataset_excludes_test_fixtures():
    files = excel_files(DATASET_DIR)

    filenames = {path.name for path in files}

    assert "90_expected_results.xlsx" not in filenames
    assert "91_error_catalog.xlsx" not in filenames


def test_phase1_dataset_contains_required_sheets():
    data = load_dataset(DATASET_DIR)

    required_sheets = {
        "01_company_plant.xlsx::company",
        "01_company_plant.xlsx::plant",
        "02_period.xlsx::period",
        "07_product_master.xlsx::product",
        "08_material_master.xlsx::material",
        "10_bom.xlsx::bom_version",
        "10_bom.xlsx::bom_item",
        "11_routing.xlsx::routing_version",
        "11_routing.xlsx::routing_operation",
        "12_standard_cost.xlsx::standard_cost",
        "12_standard_cost.xlsx::standard_cost_detail",
        "20_work_order.xlsx::work_order",
        "21_production_output.xlsx::production_output",
        "22_material_issue.xlsx::material_issue",
        "23_labor_transaction.xlsx::labor_transaction",
        "24_gl_transaction.xlsx::gl_transaction",
    }

    assert required_sheets.issubset(data.keys())


# --- Phase 2 8단계: GA 규정 정합 구조를 위한 최소 데이터 모델 추가 ---
# 아래 신규 컬럼/파일은 스키마(구조)만 준비한 것이며, 실제 GA rate/GFM
# 수치나 industry_type/company_size 등의 실제 분류값은 넣지 않았다 —
# 기존 계약 행의 신규 컬럼 값은 전부 비어 있어야 한다(None). 예외는
# plant_code 하나뿐이다 — 이 값은 "실제 GA 요율/GFM 데이터"가 아니라
# 20_work_order.xlsx(20건 전부 plant_code=PL01) +
# 01_company_plant.xlsx(회사 전체에 공장이 PL01 하나뿐)로 이미 확정되는
# 구조적 사실이라서 별도 조사 라운드에서 채워 넣었다(아래
# test_contract_plant_code_is_pl01_derived_from_work_order_and_plant_master
# 참고). fiscal_year/industry_type/company_size/contract_agreement_date는
# 여전히 채울 근거가 없어 비어 있다.

def test_contract_has_ga_regulatory_columns_but_fiscal_year_industry_company_size_still_blank():
    data = load_dataset(DATASET_DIR)
    rows = data["30_contract.xlsx::contract"]

    assert len(rows) == 3
    for r in rows:
        assert "plant_code" in r
        assert "fiscal_year" in r
        assert "industry_type" in r
        assert "company_size" in r
        assert r["fiscal_year"] is None
        assert r["industry_type"] is None
        assert r["company_size"] is None


def test_contract_plant_code_is_pl01_derived_from_work_order_and_plant_master():
    # plant_code="PL01"은 임의로 채운 추정값이 아니라, 이 데이터셋 안에서
    # 이미 확정되는 구조적 사실이다:
    #   1) 20_work_order.xlsx의 20건 전부(계약 연결 여부와 무관하게) plant_code
    #      가 "PL01" 하나뿐이다.
    #   2) 01_company_plant.xlsx의 plant 시트에는 회사(HB01) 전체에 공장이
    #      "PL01" 단 하나뿐이다 — 다른 값이 존재할 수 있는 여지 자체가 없다.
    # 이 두 사실을 함께 확인해, 이번에 넣은 값이 "그럴듯한 추정"이 아니라
    # 기존 마스터 데이터로부터 100% 확정 가능한 값이었음을 회귀로 고정한다.
    data = load_dataset(DATASET_DIR)
    contracts = data["30_contract.xlsx::contract"]
    work_orders = data["20_work_order.xlsx::work_order"]
    plants = data["01_company_plant.xlsx::plant"]

    assert len(contracts) == 3
    for r in contracts:
        assert r["plant_code"] == "PL01"

    assert len(work_orders) == 20
    assert all(w["plant_code"] == "PL01" for w in work_orders)

    assert len(plants) == 1
    assert plants[0]["plant_code"] == "PL01"


def test_contract_has_agreement_date_column_but_all_blank():
    # reference_date의 원칙적 소스(계약동의서 작성일자)를 위한 컬럼만 스키마로
    # 준비한다 — start_date 등을 대신 쓰지 않기로 한 결정에 따라, 실제 값은
    # 넣지 않고 전부 비워 둔다.
    data = load_dataset(DATASET_DIR)
    rows = data["30_contract.xlsx::contract"]

    assert len(rows) == 3
    for r in rows:
        assert "contract_agreement_date" in r
        assert r["contract_agreement_date"] is None


def test_contract_type_is_product_structure_axis_not_legal_pricing_method():
    # contract_type(MULTI_PRODUCT/SINGLE_PRODUCT/PROTOTYPE)은 제품 구성
    # 방식(다품종/단일품종/시제)을 나타내는 분류축이다. 방산원가규칙 제28조와
    # 시행세칙 제32조가 GA 기준일을 가르는 데 쓰는 "법정 계약방식"(확정계약/
    # 개산계약/중도확정계약/특정비목불확정계약)과는 서로 다른 축이며, 이
    # 사실을 조사에서 확인했다.
    #
    # 이 테스트는 딱 두 가지만 검증한다(전체 컬럼 집합은 고정하지 않는다 —
    # 법정 계약방식과 무관한 정상적인 스키마 확장까지 실패시키는 것은 이
    # 테스트의 목적을 벗어난다):
    #   1) 현재 관측되는 contract_type 값이 제품구성 축의 값들뿐이라는 것.
    #   2) 신규 GA 규정 정합 조회 함수(resolve_ga_actual_rate/
    #      resolve_ga_ceiling_rate)가 contract_type을 매칭 키로 읽지
    #      않는다는 것 — contract_type을 법정 계약방식인 것처럼 rate
    #      선택에 사용하는 코드가 생기면 이 조건이 깨져 테스트가 실패한다.
    #      (레거시 _resolve_ga_rate()는 원래부터 contract_type을 쓰는
    #      별도 축이라 이 검사 대상이 아니다 — 보호 대상이며 건드리지 않음.)
    data = load_dataset(DATASET_DIR)
    rows = data["30_contract.xlsx::contract"]

    assert len(rows) == 3
    observed_contract_types = {r["contract_type"] for r in rows}
    assert observed_contract_types == {"MULTI_PRODUCT", "SINGLE_PRODUCT", "PROTOTYPE"}

    import inspect
    from manufacturing_cost_engine import cost_engine as ce

    for fn in (ce.resolve_ga_actual_rate, ce.resolve_ga_ceiling_rate):
        source = inspect.getsource(fn)
        assert '"contract_type"' not in source


def test_material_issue_has_supply_type_column_but_all_blank():
    data = load_dataset(DATASET_DIR)
    rows = data["22_material_issue.xlsx::material_issue"]

    assert len(rows) == 62
    for r in rows:
        assert "supply_type" in r
        assert r["supply_type"] is None


def test_cost_rate_rule_file_exists_with_expected_columns_and_zero_rows():
    data = load_dataset(DATASET_DIR)
    key = "32_cost_rate_rule.xlsx::cost_rate_rule"

    assert key in data
    rows = data[key]
    assert rows == []


def test_material_issue_dup_file_still_matches_material_issue_after_schema_change():
    # 22_material_issue.xlsx에 supply_type 컬럼을 추가하면서, 바이트 단위로
    # 동일해야 하는 25_material_issue_dup.xlsx도 함께 갱신했다 — 두 파일이
    # 계속 동일한 내용을 담고 있는지 확인한다(중복탐지 테스트의 전제 보존).
    from manufacturing_cost_engine.loader import duplicate_file_groups

    groups = duplicate_file_groups(DATASET_DIR)
    assert len(groups) == 1
    _, names = groups[0]
    assert names == ["22_material_issue.xlsx", "25_material_issue_dup.xlsx"]


# --- Phase 2 9단계: Budget GFM grain(contract x material_code) 데이터 모델 ---
# contract_material_supply_type.xlsx는 헤더만 준비한 신규 파일이다. 실제
# GOVERNMENT/COMPANY 분류값은 어떤 계약·자재에도 입력하지 않았다 — 데이터
# 행이 0건이라는 사실 자체가 그 확인이다.

def test_contract_material_supply_type_file_exists_with_expected_columns_and_zero_rows():
    data = load_dataset(DATASET_DIR)
    key = "contract_material_supply_type.xlsx::contract_material_supply_type"

    assert key in data
    rows = data[key]
    assert rows == []


def test_contract_material_supply_type_key_columns_match_existing_master_naming():
    # contract_no/material_code라는 컬럼명이 기존 마스터(30_contract.xlsx,
    # 08_material_master.xlsx)의 동일한 키 컬럼명과 정확히 일치하는지
    # 확인한다 — grain(계약 x 자재) 조인이 향후 이름 불일치 없이 가능해야
    # 한다는 전제를 미리 검증해 둔다.
    data = load_dataset(DATASET_DIR)

    contracts = data["30_contract.xlsx::contract"]
    materials = data["08_material_master.xlsx::material"]

    assert contracts and "contract_no" in contracts[0]
    assert materials and "material_code" in materials[0]

    import openpyxl
    wb = openpyxl.load_workbook(DATASET_DIR / "contract_material_supply_type.xlsx")
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == ["contract_no", "material_code", "supply_type"]


# --- E-013/E-014/E-030 데이터 정정값 회귀 고정 ---
# 아래 2개 테스트는 특정 셀 값을 "그냥 이 값이니까" 고정하는 것이 아니다.
# phase1_dataset_build_spec.md가 명시한 조건(E-014: posting_date가 2026-06월,
# E-030: 이 행은 "1행만 의도적 불일치" 규정 밖이라 산술이 맞아야 함)을 만족시키기
# 위해 선택한 데이터 정정값이며, 그 조건이 다시 깨지지 않게 잠가 둔다.

def test_gl_transaction_e014_row_posting_date_satisfies_closed_period_condition():
    # build_spec.md:923 "posting_date가 2026-06인 행 1건(마감 기간, E-014)".
    # GL-260731-021 line 2는 이 조건을 만족시키기 위해 posting_date를
    # 2026-06-01로 정정했다(원본 저자가 실제로 쓴 값이라는 근거는 없음 —
    # 2026-06월 조건을 만족하는 대체값). 이 행의 period_key는 처음부터
    # "2026-06"으로 되어 있었으므로, posting_date만 그 기간과 일치시키면
    # validate_gl_period()가 PERIOD_CLOSED를 정확히 1건 발생시킨다(E-014).
    # posting_date가 다시 2026-07대로 되돌아가면 PERIOD_MISMATCH만 남고
    # PERIOD_CLOSED는 사라지므로, 이 회귀는 그 되돌림을 잡아낸다.
    data = load_dataset(DATASET_DIR)
    rows = data["24_gl_transaction.xlsx::gl_transaction"]

    target = [
        r for r in rows
        if r.get("document_no") == "GL-260731-021" and r.get("line_no") == 2
    ]
    assert len(target) == 1
    row = target[0]

    assert row["period_key"] == "2026-06"
    assert row["posting_date"] == "2026-06-01"


def test_labor_transaction_e030_sibling_row_hours_are_internally_consistent():
    # phase1_dataset_build_spec.md:882 "actual_hours = regular_hours +
    # overtime_hours — 1행만 의도적 불일치(E-030)". LB-2607-045는 그
    # "1행"이 아니다(그 자리는 LB-2607-047) — 이 행은 E-028(NEGATIVE_OVERTIME,
    # overtime_hours=-2)의 지정 행일 뿐이므로 산술은 맞아야 한다.
    # actual_hours를 -1.5로 정정한 것은 "regular_hours(0.5) + overtime_hours(-2)
    # = -1.5"를 맞추는 값이며, 원본 저자가 실제로 입력한 값이라는 근거는 없다
    # (조건 충족값). 이 값이 다시 산술과 어긋나면(예: 이전 값 -1로 되돌아가면)
    # validate_labor()가 이 행에 대해서도 HOURS_SUM_MISMATCH를 추가로
    # 발생시켜 E-030이 1건이 아니라 2건이 되므로, 이 회귀는 그 상태를 잡아낸다.
    data = load_dataset(DATASET_DIR)
    rows = data["23_labor_transaction.xlsx::labor_transaction"]

    target = [r for r in rows if r.get("labor_doc_no") == "LB-2607-045"]
    assert len(target) == 1
    row = target[0]

    assert row["regular_hours"] == 0.5
    assert row["overtime_hours"] == -2
    assert row["actual_hours"] == -1.5
    assert row["actual_hours"] == row["regular_hours"] + row["overtime_hours"]


def test_contract_material_supply_type_no_real_government_or_company_values():
    # 실제 관급/사급 분류값이 어떤 형태로도 들어있지 않은지 최종 확인한다.
    data = load_dataset(DATASET_DIR)
    rows = data["contract_material_supply_type.xlsx::contract_material_supply_type"]

    assert len(rows) == 0
    assert all(r.get("supply_type") in (None,) for r in rows)


# --- Phase 1 regression lock (files/phase1_validation_baseline.md, commit 2d6afc8) ---
#
# 아래 섹션은 Phase 1을 "고치는" 것이 아니라, 현재 실제 데이터셋에 대한
# validation/원가계산 결과를 회귀 기준선으로 잠그는 것이다. cli.py의
# main()과 정확히 동일한 순서로 검증 함수를 호출해 issues 리스트를
# 재구성한다(cli.py의 조립 로직 자체는 별도 함수로 분리되어 있지 않아
# 재사용할 수 없으므로, 여기서는 그 순서를 그대로 따라간다 — 새 검증
# 함수를 만들지 않는다).

def _build_phase1_issues(data):
    def rows(file, sheet):
        return data.get(f"{file}::{sheet}", [])

    issues = []
    issues += validate_duplicate_files(duplicate_file_groups(DATASET_DIR))
    issues += validate_period_rows(rows("02_period.xlsx", "period"))

    products = rows("07_product_master.xlsx", "product")
    issues += validate_work_orders(rows("20_work_order.xlsx", "work_order"), products)

    contracts = rows("30_contract.xlsx", "contract")
    issues += validate_contract(contracts, rows("20_work_order.xlsx", "work_order"))

    direct_expenses = rows("31_direct_expense.xlsx", "direct_expense")
    issues += validate_direct_expense(
        direct_expenses, contracts, rows("20_work_order.xlsx", "work_order")
    )

    issues += validate_material_issues(
        rows("22_material_issue.xlsx", "material_issue"),
        rows("08_material_master.xlsx", "material"),
        rows("20_work_order.xlsx", "work_order"),
        rows("03_cost_center.xlsx", "cost_center"),
    )

    issues += validate_bom_issues(
        rows("20_work_order.xlsx", "work_order"),
        rows("10_bom.xlsx", "bom_item"),
        rows("22_material_issue.xlsx", "material_issue"),
        rows("08_material_master.xlsx", "material"),
        rows("06_uom_conversion.xlsx", "uom_conversion"),
    )

    issues += validate_bom_version(rows("10_bom.xlsx", "bom_version"))
    issues += validate_tolerance_rules(rows("14_tolerance_rule.xlsx", "tolerance_rule"))
    issues += validate_routing(
        rows("20_work_order.xlsx", "work_order"),
        rows("11_routing.xlsx", "routing_version"),
        rows("11_routing.xlsx", "routing_operation"),
        rows("09_work_center.xlsx", "work_center"),
    )

    wo_map = {r.get("wo_no"): r for r in rows("20_work_order.xlsx", "work_order")}
    routing_index = {}
    for r in rows("11_routing.xlsx", "routing_operation"):
        routing_index[(r.get("routing_version_id"), r.get("operation_seq"))] = r
    routing_version_by_product = {}
    for h in rows("11_routing.xlsx", "routing_version"):
        product_code = h.get("product_code")
        routing_version_id = h.get("routing_version_id")
        if product_code is not None and routing_version_id is not None:
            routing_version_by_product.setdefault(product_code, routing_version_id)

    def resolve_routing_version_id(wo_row):
        explicit = wo_row.get("routing_version_id")
        if explicit is not None:
            return explicit
        return routing_version_by_product.get(wo_row.get("product_code"))

    labor_rows = rows("23_labor_transaction.xlsx", "labor_transaction")
    labor_index = {}
    for r in labor_rows:
        wo = r.get("wo_no")
        wo_row = wo_map.get(wo)
        if not wo_row:
            continue
        routing_version_id = resolve_routing_version_id(wo_row)
        if routing_version_id is None:
            continue
        routing_row = routing_index.get((routing_version_id, r.get("operation_seq")))
        if routing_row is not None:
            labor_index[(wo, r.get("operation_seq"))] = routing_row
    issues += validate_labor(labor_rows, labor_index)
    issues += validate_labor_hours(
        rows("20_work_order.xlsx", "work_order"),
        labor_rows,
        rows("21_production_output.xlsx", "production_output"),
        rows("11_routing.xlsx", "routing_operation"),
        rows("11_routing.xlsx", "routing_version"),
    )
    issues += validate_gl_balance(rows("24_gl_transaction.xlsx", "gl_transaction"))
    issues += validate_gl_period(
        rows("24_gl_transaction.xlsx", "gl_transaction"),
        rows("02_period.xlsx", "period"),
    )
    issues += validate_account_mapping(
        rows("24_gl_transaction.xlsx", "gl_transaction"),
        rows("05_account_mapping.xlsx", "account_mapping"),
    )
    issues += validate_standard_cost(
        rows("12_standard_cost.xlsx", "standard_cost"),
        rows("12_standard_cost.xlsx", "standard_cost_detail"),
        products,
        rows("04_cost_element.xlsx", "cost_element"),
        rows("08_material_master.xlsx", "material"),
    )

    work_orders = rows("20_work_order.xlsx", "work_order")
    production_outputs = rows("21_production_output.xlsx", "production_output")
    work_centers = rows("09_work_center.xlsx", "work_center")
    overhead_rates = rows("13_overhead_rate.xlsx", "overhead_rate")

    issues += validate_actual_cost(
        work_orders, production_outputs, labor_rows, work_centers, overhead_rates,
    )
    issues += validate_gl_reconciliation(
        rows("24_gl_transaction.xlsx", "gl_transaction"),
        rows("05_account_mapping.xlsx", "account_mapping"),
        work_orders,
        rows("22_material_issue.xlsx", "material_issue"),
        labor_rows,
        rows("08_material_master.xlsx", "material"),
        products,
        rows("02_period.xlsx", "period"),
    )
    return issues


def _load_expected_results_rows():
    # 90_expected_results.xlsx는 검증 픽스처라 load_dataset()의 로딩 대상에서
    # 제외되므로(test_excel_files_exclude_expected_result_files 참고) openpyxl로
    # 직접 읽는다.
    import openpyxl
    wb = openpyxl.load_workbook(DATASET_DIR / "90_expected_results.xlsx", read_only=True)
    ws = wb["validation_summary"]
    rows = list(ws.iter_rows(values_only=True))
    return rows[1:]  # (assertion_id, expected_status, expected_value, note)


# 현재 데이터셋 전체에서 실제로 발생하는 validation code별 건수(총 64건, 30개
# 코드). files/phase1_validation_baseline.md §2에 기록된 상태와 동일하다.
EXPECTED_ISSUE_CODE_COUNTS = {
    "UOM_CONVERSION_MISSING": 14,
    "MISSING_MATERIAL_ISSUE": 13,
    "OPERATION_CODE_MISMATCH": 6,
    "NOT_IN_BOM": 3,
    "GL_RECON_DIFFERENCE": 2,
    "EXCLUDED_WO": 2,
    "DUPLICATE_FILE": 1,
    "UNKNOWN_PRODUCT": 1,
    "INVALID_DECIMAL": 1,
    "UNKNOWN_MATERIAL": 1,
    "UNKNOWN_WO": 1,
    "NEGATIVE_QUANTITY": 1,
    "BOM_OVER_ISSUE": 1,
    "TOLERANCE_AMBIGUOUS": 1,
    "UNKNOWN_ROUTING": 1,
    "UNKNOWN_ROUTING_OPERATION": 1,
    "INVALID_LABOR_RATE": 1,
    "NEGATIVE_OVERTIME": 1,
    "HOURS_SUM_MISMATCH": 1,
    "EXCESSIVE_LABOR_HOURS": 1,
    "GL_UNBALANCED_DOCUMENT": 1,
    "PERIOD_MISMATCH": 1,
    "PERIOD_CLOSED": 1,
    "UNMAPPED_GL": 1,
    "MAPPING_AMBIGUOUS": 1,
    "STD_DETAIL_SUM_MISMATCH": 1,
    "STANDARD_COST_MISSING": 1,
    "REWORK_REVIEW_REQUIRED": 1,
    "ZERO_DENOMINATOR": 1,
    "NO_PRODUCTION_OUTPUT": 1,
}

# These seven assertions are known unresolved Phase 1 dataset/fixture
# discrepancies documented in files/phase1_validation_baseline.md. They are
# intentionally not normalized or repaired by the regression test. This dict
# is the single source of truth for their currently-observed actual values —
# it is not a claim that these values are correct, only that they are the
# current, unchanged baseline.
KNOWN_UNRESOLVED_ACTUAL_VALUES = {
    "E-002": 0,
    "E-003": 3,
    "E-011": 0,
    "E-016": 14,
    "E-018": 13,
    "E-024": 0,
    "E-027": 2,
}


def test_phase1_validation_issue_counts_are_stable():
    # 실제 Phase 1 데이터셋 전체에 대한 validation issue 총건수/코드별 건수가
    # files/phase1_validation_baseline.md에 기록된 상태에서 벗어나지 않는지
    # 잠근다. 코드 분류(예: BOM 4분류) 자체는 synthetic 테스트가 검증하지만,
    # "실제 20개 WO/62개 material_issue 등에 대해 정확히 몇 건이 나오는가"는
    # 지금까지 어떤 pytest도 보호하지 않았다.
    data = load_dataset(DATASET_DIR)
    issues = _build_phase1_issues(data)

    assert len(issues) == 64
    assert Counter(i.code for i in issues) == Counter(EXPECTED_ISSUE_CODE_COUNTS)


def test_phase1_expected_results_baseline_is_stable():
    # 90_expected_results.xlsx의 32개 assertion 중 25개는 실제 결과와
    # 일치하고, 7개(KNOWN_UNRESOLVED_ACTUAL_VALUES)는 근거 부족/설계 문서
    # 충돌/fixture 결함으로 아직 해결되지 않은 상태다
    # (files/phase1_validation_baseline.md §3~4 참고). 이 테스트는 그
    # 25/7 구성이 그대로 유지되는지 잠글 뿐, 7개를 32/32로 맞추지 않는다.
    assert set(KNOWN_UNRESOLVED_ACTUAL_VALUES) == {
        "E-002", "E-003", "E-011", "E-016", "E-018", "E-024", "E-027",
    }

    data = load_dataset(DATASET_DIR)
    issues = _build_phase1_issues(data)
    actual_counts = Counter(i.code for i in issues)

    expected_rows = _load_expected_results_rows()
    assert len(expected_rows) == 32

    match_count = 0
    unresolved_seen = set()
    for assertion_id, expected_status, expected_value, _note in expected_rows:
        actual_value = actual_counts.get(expected_status, 0)
        if assertion_id in KNOWN_UNRESOLVED_ACTUAL_VALUES:
            unresolved_seen.add(assertion_id)
            assert actual_value == KNOWN_UNRESOLVED_ACTUAL_VALUES[assertion_id], (
                f"{assertion_id}({expected_status}): known-unresolved actual "
                f"value changed — investigate before updating this baseline"
            )
        else:
            assert actual_value == int(expected_value), (
                f"{assertion_id}({expected_status}): previously-matching "
                f"assertion regressed"
            )
            match_count += 1

    assert unresolved_seen == set(KNOWN_UNRESOLVED_ACTUAL_VALUES)
    assert match_count == 25
    assert len(unresolved_seen) == 7


def test_phase1_total_actual_cost_is_stable():
    # 전체 20개 WO에 대한 실적 원가 합계(CLI의 "Total Actual Cost (all WOs)")를
    # 잠근다. Contract 단위 합계(test_real_dataset_contract_00*)는 이미
    # test_cost_engine.py에 있으므로 중복 작성하지 않는다 — 여기서는 그
    # 테스트들이 다루지 않는 "전체 WO 합계" 하나만 본다.
    data = load_dataset(DATASET_DIR)

    def rows(file, sheet):
        return data.get(f"{file}::{sheet}", [])

    actual_cost_by_wo = calculate_actual_total_cost_by_wo(
        rows("20_work_order.xlsx", "work_order"),
        rows("22_material_issue.xlsx", "material_issue"),
        rows("23_labor_transaction.xlsx", "labor_transaction"),
        rows("08_material_master.xlsx", "material"),
        rows("09_work_center.xlsx", "work_center"),
        rows("13_overhead_rate.xlsx", "overhead_rate"),
        rows("07_product_master.xlsx", "product"),
    )
    total_actual_cost = sum(
        (v["total_cost"] for v in actual_cost_by_wo.values()), start=Decimal("0")
    )

    assert total_actual_cost == Decimal("4921044.89")


def test_phase1_wo_total_variance_is_stable():
    # 실제 데이터셋의 WO별 total_variance(CLI "Total Variance calculated for
    # 17 work order(s)")를 하나의 매핑으로 잠근다. WO-2607-009/017/020은
    # 표준원가/생산실적 부재로 계산 불가라 이 결과에 아예 나타나지 않는다
    # (이 자체도 baseline의 일부이므로 dict 비교로 함께 고정된다).
    data = load_dataset(DATASET_DIR)

    def rows(file, sheet):
        return data.get(f"{file}::{sheet}", [])

    variance_by_wo = calculate_total_variance_by_wo(
        rows("20_work_order.xlsx", "work_order"),
        rows("22_material_issue.xlsx", "material_issue"),
        rows("23_labor_transaction.xlsx", "labor_transaction"),
        rows("08_material_master.xlsx", "material"),
        rows("09_work_center.xlsx", "work_center"),
        rows("13_overhead_rate.xlsx", "overhead_rate"),
        rows("07_product_master.xlsx", "product"),
        rows("21_production_output.xlsx", "production_output"),
        rows("12_standard_cost.xlsx", "standard_cost"),
    )

    expected_total_variance = {
        "WO-2607-001": Decimal("318002.00"),
        "WO-2607-002": Decimal("-795256.00"),
        "WO-2607-003": Decimal("-1077685.12"),
        "WO-2607-004": Decimal("-457780.00"),
        "WO-2607-005": Decimal("-557520.00"),
        "WO-2607-006": Decimal("-795256.00"),
        "WO-2607-007": Decimal("-1078012.80"),
        "WO-2607-008": Decimal("-212872.38"),
        "WO-2607-010": Decimal("-657440.00"),
        "WO-2607-011": Decimal("-1078012.80"),
        "WO-2607-012": Decimal("-562440.00"),
        "WO-2607-013": Decimal("-557520.00"),
        "WO-2607-014": Decimal("-942261.20"),
        "WO-2607-015": Decimal("-364796.81"),
        "WO-2607-016": Decimal("-556040.00"),
        "WO-2607-018": Decimal("-433538.00"),
        "WO-2607-019": Decimal("-1162624.00"),
    }

    assert len(variance_by_wo) == 17
    actual_total_variance = {
        wo_no: v["total_variance"] for wo_no, v in variance_by_wo.items()
    }
    assert actual_total_variance == expected_total_variance
