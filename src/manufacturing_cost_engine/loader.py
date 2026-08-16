from pathlib import Path

from openpyxl import load_workbook


def load_excel_file(file_path: str | Path) -> dict[str, list[dict]]:
    """
    Load an Excel workbook.

    Returns:
        {
            "sheet_name": [
                {"column": value, ...},
                ...
            ]
        }
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    workbook = load_workbook(
        filename=path,
        data_only=True,
    )

    result = {}

    for worksheet in workbook.worksheets:
        rows = list(
            worksheet.iter_rows(
                values_only=True
            )
        )

        if not rows:
            result[worksheet.title] = []
            continue

        headers = [
            str(value).strip() if value is not None else ""
            for value in rows[0]
        ]

        if any(not header for header in headers):
            raise ValueError(
                f"Empty header detected: "
                f"{path.name} / {worksheet.title}"
            )

        data = []

        for row_number, row in enumerate(rows[1:], start=2):
            if all(value is None for value in row):
                continue

            record = {
                header: value
                for header, value in zip(headers, row)
            }

            record["_source_row"] = row_number
            record["_source_sheet_name"] = worksheet.title

            data.append(record)

        result[worksheet.title] = data

    return result
